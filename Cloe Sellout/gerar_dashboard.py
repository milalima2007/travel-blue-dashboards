#!/usr/bin/env python3
"""
gerar_dashboard.py — Dashboard Cronogramas Cloe
Le los 3 archivos Excel locales y genera dashboard.html con todos los datos embebidos.
No necesita internet ni Google Sheets.

Uso: python gerar_dashboard.py
"""
import os
import json
import re
import sys
from datetime import datetime, date

try:
    from openpyxl import load_workbook
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import load_workbook

def show_error(msg):
    """Muestra error visible tanto en consola como en ventana (cuando es .exe)."""
    print("ERROR:", msg)
    try:
        import ctypes
        ctypes.windll.user32.MessageBoxW(0, str(msg), "Dashboard Cloe - Error", 0x10)
    except Exception:
        pass
    input("Presiona Enter para cerrar...")
    sys.exit(1)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT   = os.path.join(BASE_DIR, "dashboard.html")

# ── Configuración de archivos Excel ──────────────────────────────────────────
SECTIONS_CFG = [
    {
        "key":              "nuevos",
        "label":            "Nuevos Desarrollos",
        "file":             "Cronogramas_Productos_Nuevos_Cloe_fechas_rev_ok.xlsx",
        "use_comite_start": True,
        "sku_from_sheet":   True,   # nombre de la pestaña = SKU
    },
    {
        "key":              "promocional",
        "label":            "Promocional",
        "file":             "Cronograma Promocional Cloe.xlsx",
        "use_comite_start": False,
        "sku_from_sheet":   False,
    },
    {
        "key":              "peanuts",
        "label":            "Peanuts",
        "file":             "Cronograma Peanuts Cloe.xlsx",
        "use_comite_start": False,
        "sku_from_sheet":   False,
    },
]

# ── Helpers ───────────────────────────────────────────────────────────────────
def fmt_date(v):
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    if v and re.match(r"\d{4}-\d{2}-\d{2}", str(v)):
        return str(v)[:10]
    return None

def num_or_null(v):
    if v is None or str(v).strip() == "":
        return None
    try:
        return int(float(str(v)))
    except (ValueError, TypeError):
        return None

def clean_str(v):
    if v is None:
        return ""
    return str(v).strip()

VALID_STATUS = {"Completada", "En curso", "Pendiente"}

def read_section(cfg):
    path = os.path.join(BASE_DIR, cfg["file"])
    if not os.path.exists(path):
        print(f"  ADVERTENCIA: no encontrado -> {cfg['file']}")
        return []

    wb = load_workbook(path, data_only=True)
    products = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 6:
            continue

        # Fila 2 (índice 2): nombre del producto en columna 0
        product_name = clean_str(rows[2][0] if len(rows[2]) > 0 else "")
        if not product_name or "CRONOGRAMA" in product_name.upper():
            continue

        sku = sheet_name if cfg["sku_from_sheet"] else re.sub(r"[^A-Z0-9]", "-", sheet_name.upper())

        tasks = []
        for row in rows[5:]:
            if not row or row[0] is None:
                continue
            num = num_or_null(row[0])
            if num is None:
                continue
            tarea = clean_str(row[1] if len(row) > 1 else None)
            if not tarea:
                continue

            status_raw = clean_str(row[7] if len(row) > 7 else None)
            status = status_raw if status_raw in VALID_STATUS else "Pendiente"

            tasks.append({
                "num":           num,
                "name":          tarea,
                "duration":      num_or_null(row[2] if len(row) > 2 else None),
                "depends_on":    clean_str(row[3] if len(row) > 3 else None) or None,
                "dep_type":      clean_str(row[4] if len(row) > 4 else None) or None,
                "plan_start":    fmt_date(row[5] if len(row) > 5 else None),
                "plan_end":      fmt_date(row[6] if len(row) > 6 else None),
                "status":        status,
                "real_start":    fmt_date(row[8] if len(row) > 8 else None),
                "real_end":      fmt_date(row[9] if len(row) > 9 else None),
                "real_duration": num_or_null(row[10] if len(row) > 10 else None),
                "deviation":     num_or_null(row[11] if len(row) > 11 else None),
                "comments":      clean_str(row[12] if len(row) > 12 else None),
            })

        if tasks:
            products.append({"sku": sku, "title": product_name, "tasks": tasks})
            print(f"    {sku}: {len(tasks)} tareas")

    return products


def build_sections_data():
    sections_data = {}
    for cfg in SECTIONS_CFG:
        print(f"  Leyendo {cfg['file']}...")
        products = read_section(cfg)
        sections_data[cfg["key"]] = {
            "label":          cfg["label"],
            "useComiteStart": cfg["use_comite_start"],
            "products":       products,
        }
    return sections_data


def build_html(sections_data):
    generated_at = datetime.now().strftime("%d/%m/%Y %H:%M")
    sections_json = json.dumps(sections_data, ensure_ascii=False)
    html = HTML_TEMPLATE.replace("__SECTIONS_DATA__", sections_json)
    html = html.replace("__GENERATED_AT__", generated_at)
    return html


HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<meta name="robots" content="noindex, nofollow">
<title>Dashboard Cronogramas · Cloe</title>
<style>
:root{--bg:#0a0e1a;--card:#111827;--border:#1e293b;--text:#e2e8f0;--dim:#64748b;
  --accent:#f59e0b;--green:#10b981;--red:#ef4444;--blue:#3b82f6;--purple:#8b5cf6;}
*{margin:0;padding:0;box-sizing:border-box}
body{background:var(--bg);color:var(--text);font-family:'Segoe UI',system-ui,sans-serif;font-size:13px;line-height:1.5}
.header{background:linear-gradient(135deg,#111827 0%,#1e293b 100%);padding:14px 24px;border-bottom:2px solid var(--accent);display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:8px;}
.logo{display:flex;align-items:center;gap:10px;}
.logo-text .brand{font-size:17px;font-weight:700;color:#fff;display:block;}
.logo-text .sub{font-size:11px;color:var(--dim);}
.header-right{text-align:right;font-size:11px;color:var(--dim);}
.section-tabs{display:flex;gap:6px;padding:14px 24px 0;flex-wrap:wrap;}
.section-tab{padding:8px 18px;border-radius:8px 8px 0 0;font-size:12px;font-weight:700;cursor:pointer;background:#1e293b;border:1px solid var(--border);border-bottom:none;color:var(--dim);transition:all .15s;}
.section-tab:hover{color:var(--text);}
.section-tab.active{background:var(--card);color:var(--accent);border-color:var(--accent);}
.kpi-row{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;padding:16px 24px;}
.kpi{background:var(--card);border:1px solid var(--border);border-radius:10px;padding:14px 18px;text-align:center;}
.kpi .lbl{font-size:10px;color:var(--dim);text-transform:uppercase;letter-spacing:.5px;}
.kpi .val{font-size:20px;font-weight:800;color:var(--accent);margin:4px 0;}
.kpi .cnt{font-size:11px;color:var(--dim);}
.content{padding:16px 24px;max-width:1600px;margin:0 auto;}
.card{background:var(--card);border:1px solid var(--border);border-radius:10px;padding:18px;margin-bottom:16px;}
.section-hdr{display:flex;align-items:center;justify-content:space-between;gap:8px;margin-bottom:10px;flex-wrap:wrap;}
.section-hdr h2{font-size:13px;font-weight:700;color:var(--accent);text-transform:uppercase;letter-spacing:.4px;}
.hdr-actions{display:flex;gap:6px;flex-wrap:wrap;}
.ptabs{display:flex;gap:6px;}
.ptab,.export-btn{padding:5px 14px;border-radius:6px;font-size:11px;font-weight:600;cursor:pointer;background:#1e293b;border:1px solid var(--border);color:var(--dim);transition:all .15s;}
.ptab:hover,.export-btn:hover{color:var(--text);border-color:var(--dim);}
.ptab.active{background:#f59e0b22;color:var(--accent);border-color:var(--accent);}
.tbl-wrap{overflow-x:auto;}
table.dt{width:100%;border-collapse:collapse;font-size:11px;}
table.dt th{text-align:left;padding:5px 8px;border-bottom:2px solid var(--accent);color:var(--accent);font-weight:700;font-size:10px;text-transform:uppercase;white-space:nowrap;}
table.dt td{padding:6px 8px;border-bottom:1px solid var(--border);white-space:nowrap;}
table.dt tr.product-row{cursor:pointer;}
table.dt tr.product-row:hover td{background:#1e293b55;}
table.dt tr.product-row.active td{background:#f59e0b18;}
.bar-track{background:#1e293b;border-radius:4px;height:8px;width:140px;overflow:hidden;display:inline-block;vertical-align:middle;}
.bar-fill{height:100%;border-radius:4px;background:var(--green);}
.badge-p{display:inline-block;padding:2px 8px;border-radius:20px;font-size:11px;font-weight:700;}
.status-pendiente{background:#64748b22;color:var(--dim);}
.status-en-curso{background:#f59e0b22;color:var(--accent);}
.status-completada{background:#10b98122;color:var(--green);}
.deviation-ok{color:var(--green);font-weight:700;}
.deviation-warn{color:var(--red);font-weight:700;}
.detail-title{font-size:15px;font-weight:800;color:var(--text);}
.detail-sub{font-size:11px;color:var(--dim);margin-top:2px;}
.dtable{width:100%;border-collapse:collapse;font-size:11.5px;}
.dtable th{background:#1e293b;color:var(--dim);font-weight:600;padding:5px 8px;text-align:left;border-bottom:1px solid var(--border);white-space:nowrap;}
.dtable td{padding:5px 8px;border-bottom:1px solid rgba(255,255,255,.04);color:var(--text);vertical-align:middle;}
.dtable tbody tr:hover{background:rgba(255,255,255,.03);}
.legend{display:flex;gap:16px;flex-wrap:wrap;margin-bottom:14px;font-size:11px;color:var(--dim);}
.legend-item{display:flex;align-items:center;gap:6px;}
.legend-dot{width:10px;height:10px;border-radius:3px;display:inline-block;}
.gantt{display:flex;flex-direction:column;gap:0;}
.gantt-group{display:grid;grid-template-columns:26px 250px 1fr 70px;gap:0 10px;align-items:stretch;padding:6px 0;border-bottom:1px solid rgba(255,255,255,.04);}
.gantt-group:last-child{border-bottom:none;}
.gantt-num{font-size:11px;color:var(--dim);text-align:right;padding-top:3px;}
.gantt-name-col{display:flex;flex-direction:column;justify-content:center;}
.gantt-task{font-size:12px;font-weight:500;color:var(--text);line-height:1.3;}
.gantt-comment-inline{font-size:10px;color:var(--dim);font-style:italic;margin-top:2px;}
.gantt-bars-col{display:flex;flex-direction:column;gap:3px;justify-content:center;}
.gantt-bar-row{display:flex;align-items:center;gap:6px;}
.gantt-bar-label{font-size:10px;color:var(--dim);width:28px;text-align:right;flex-shrink:0;}
.gantt-bar-label.real{color:#94a3b8;}
.gantt-track-wrap{position:relative;height:14px;background:#1e293b;border-radius:4px;overflow:hidden;flex:1;}
.gantt-bar{position:absolute;top:2px;bottom:2px;border-radius:3px;opacity:.92;}
.gantt-bar-real-solid{position:absolute;top:2px;bottom:2px;border-radius:3px;background:rgba(148,163,184,0.55);}
.gantt-today{position:absolute;top:-2px;bottom:-2px;width:2px;background:var(--accent);}
.gantt-dev-col{display:flex;flex-direction:column;justify-content:center;gap:3px;}
.gantt-dev{font-size:11px;text-align:right;}
.gantt-dev-spacer{height:14px;}
.gantt-axis{margin-bottom:6px;padding-bottom:6px;border-bottom:1px solid var(--border);}
.gantt-axis-track{position:relative;height:14px;}
.gantt-month-tick{position:absolute;top:0;font-size:10px;color:var(--dim);font-weight:700;white-space:nowrap;transform:translateX(-2px);border-left:1px solid var(--border);padding-left:4px;}
footer{text-align:center;color:var(--dim);font-size:11px;padding:16px;border-top:1px solid var(--border);margin-top:8px;}
@media print{
  :root{--bg:#fff;--card:#fff;--border:#ccc;--text:#111;--dim:#555;--accent:#92400e;--green:#15803d;--red:#b91c1c;}
  .header,.section-tabs,.hdr-actions,.ptabs,footer{display:none !important;}
  body{background:#fff;color:#111;font-size:11px;}
  .content{padding:0;max-width:none;}
  .card{break-inside:avoid;border:1px solid #ccc;box-shadow:none;}
  .bar-track,.gantt-track-wrap{background:#eee;}
  .badge-p{border:1px solid #ccc;}
  .dtable th{background:#f3f4f6;color:#555;}
  .dtable td{color:#111;border-bottom:1px solid #ddd;}
}
@media(max-width:768px){
  .kpi-row{grid-template-columns:repeat(2,1fr);}
  .gantt-row{grid-template-columns:22px 1fr;grid-template-areas:"num task" "bar bar" "dev dev";}
  .gantt-task{grid-area:task;} .gantt-num{grid-area:num;}
  .gantt-track-wrap{grid-area:bar;} .gantt-dev{grid-area:dev;text-align:left;}
}
</style>
</head>
<body>

<div class="header">
  <div class="logo">
    <svg viewBox="0 0 48 48" width="40" height="40" xmlns="http://www.w3.org/2000/svg">
      <rect width="48" height="48" rx="8" fill="#1B2A6B"/>
      <path d="M10 34 Q16 10 24 24 Q32 38 38 14" stroke="#FFC72C" stroke-width="3.5" fill="none" stroke-linecap="round"/>
      <circle cx="38" cy="14" r="2.5" fill="#FFC72C"/>
    </svg>
    <div class="logo-text">
      <span class="brand">Travel Blue</span>
      <span class="sub" id="section-subtitle">Cronogramas Cloe</span>
    </div>
  </div>
  <div class="header-right" id="header-right">Generado el __GENERATED_AT__</div>
</div>

<div class="section-tabs" id="section-tabs"></div>
<div class="kpi-row" id="summary-grid"></div>

<div class="content">
  <div class="card">
    <div class="section-hdr">
      <h2>📋 Resumen por producto</h2>
      <div class="hdr-actions">
        <div class="ptabs" id="product-tabs">
          <button class="ptab active" data-filter="active">En desarrollo</button>
          <button class="ptab" data-filter="done">Completados</button>
        </div>
        <button class="export-btn" id="btn-export-pdf">🖨️ Exportar PDF</button>
        <button class="export-btn" id="btn-export-excel">📊 Exportar Excel</button>
      </div>
    </div>
    <div class="tbl-wrap">
      <table class="dt">
        <thead><tr>
          <th>Producto (SKU)</th><th>Progreso</th><th>Tareas</th>
          <th>Inicio plan.</th><th>Fin estimado</th><th>Desvío máx.</th>
        </tr></thead>
        <tbody id="products-tbody"></tbody>
      </table>
    </div>
  </div>

  <div class="card">
    <div class="section-hdr"><h2>📅 Cronograma detallado</h2></div>
    <div class="legend">
      <span class="legend-item"><span class="legend-dot" style="background:#10b981"></span> Completada</span>
      <span class="legend-item"><span class="legend-dot" style="background:#f59e0b"></span> En curso</span>
      <span class="legend-item"><span class="legend-dot" style="background:#64748b"></span> Pendiente</span>
      <span class="legend-item"><span class="legend-dot" style="border:2px solid rgba(255,255,255,.5);background:transparent"></span> Fechas reales</span>
      <span class="legend-item"><span class="legend-dot" style="background:#f59e0b;width:2px;border-radius:0"></span> Hoy</span>
    </div>
    <div class="detail-title" id="detail-title">—</div>
    <div class="detail-sub" id="detail-sub" style="margin-bottom:14px;">—</div>
    <div id="detail-task-table" style="margin-bottom:18px;"></div>
    <div class="gantt" id="gantt-container"></div>
  </div>
</div>

<footer>Travel Blue · Dashboard de cronogramas Cloe · Última actualización: __GENERATED_AT__</footer>

<script>
const SECTIONS_DATA = __SECTIONS_DATA__;
const SECTION_KEYS  = Object.keys(SECTIONS_DATA);

let currentSection = SECTION_KEYS[0];
let PRODUCTS       = [];
let TODAY          = new Date().toISOString().slice(0, 10);
let activeSku      = null;
let productFilter  = 'active';

function escapeHtml(s) {
  return String(s == null ? '' : s).replace(/[&<>"']/g, c => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c]));
}
function validDateOrNull(v) {
  if (!v || !/^\d{4}-\d{2}-\d{2}$/.test(v)) return null;
  return v;
}
function fmtDisplay(v) {
  if (!v) return '—';
  const [y, m, d] = v.split('-');
  return `${d}/${m}/${y}`;
}
function statusColor(s) {
  if (s === 'Completada') return '#27AE60';
  if (s === 'En curso')   return '#fece08';
  return '#94A3B8';
}

function daysDiff(a, b) {
  if (!a || !b) return null;
  return Math.round((new Date(b) - new Date(a)) / 86400000);
}

function buildProducts(rawProducts, useComiteStart) {
  return rawProducts.map(p => {
    const tasks     = p.tasks;
    const total     = tasks.length;
    const completed = tasks.filter(t => t.status === 'Completada').length;
    const inProg    = tasks.filter(t => t.status === 'En curso').length;
    const planStarts = tasks.map(t => t.plan_start).filter(Boolean).sort();
    const planEnds   = tasks.map(t => t.plan_end).filter(Boolean).sort();
    const realEnds   = tasks.map(t => t.real_end).filter(Boolean).sort();
    const earliestStart = planStarts[0] || null;

    // Comité shift: días de retraso de la etapa 4 respecto a su plan original.
    // Para secciones con useComiteStart, los desvíos de las etapas posteriores
    // se muestran descontando este retraso (la responsabilidad de Cloe comienza
    // en la aprobación del Comité, no en el inicio original).
    let comiteShift = 0;
    let displayStart = earliestStart;
    if (useComiteStart) {
      const comite = tasks.find(t => t.num === 4);
      if (comite) {
        const cs = comite.real_end || comite.real_start || comite.plan_end || comite.plan_start;
        displayStart = cs || earliestStart;
        const shift = daysDiff(comite.plan_end, comite.real_end);
        comiteShift = shift !== null ? Math.max(0, shift) : 0;
      }
    }

    // Calcular desvío ajustado por etapa
    const tasksWithAdj = tasks.map(t => {
      let adjDev = t.deviation;
      if (useComiteStart && typeof t.deviation === 'number') {
        adjDev = t.num >= 4 ? Math.max(0, t.deviation - comiteShift) : 0;
      }
      return { ...t, adj_deviation: adjDev };
    });

    const adjDevs = tasksWithAdj.map(t => t.adj_deviation).filter(d => typeof d === 'number');
    const pctComplete   = total ? Math.round(100 * completed / total) : 0;
    const latestRealEnd = realEnds.slice(-1)[0] || null;
    // Fin estimado = siempre la tarea "Lanzamiento" (última tarea del cronograma)
    const lastTask      = tasks.slice().sort((a,b) => (a.num||0)-(b.num||0)).slice(-1)[0];
    const currentEndEstimate = lastTask
      ? (lastTask.real_end || lastTask.plan_end || planEnds.slice(-1)[0] || latestRealEnd)
      : (planEnds.slice(-1)[0] || latestRealEnd);
    return {
      sku: p.sku, title: p.title, tasks: tasksWithAdj,
      total_tasks: total, completed_tasks: completed, in_progress_tasks: inProg,
      pct_complete: pctComplete,
      plan_start: displayStart,
      current_end_estimate: currentEndEstimate,
      max_deviation: adjDevs.length ? Math.max(...adjDevs) : 0,
      comite_shift: comiteShift,
    };
  });
}

function loadSection(key) {
  currentSection = key;
  activeSku      = null;
  productFilter  = 'active';
  document.querySelectorAll('.section-tab').forEach(b => b.classList.toggle('active', b.dataset.key === key));
  document.querySelectorAll('.ptab').forEach(b => b.classList.toggle('active', b.dataset.filter === 'active'));
  const sec = SECTIONS_DATA[key];
  document.getElementById('section-subtitle').textContent = 'Cronograma · ' + sec.label + ' Cloe';
  PRODUCTS = buildProducts(sec.products, !!sec.useComiteStart);
  document.getElementById('header-right').textContent =
    PRODUCTS.length + ' productos · Generado el __GENERATED_AT__';
  renderAll();
}

function renderSectionTabs() {
  document.getElementById('section-tabs').innerHTML = SECTION_KEYS.map(k =>
    `<button class="section-tab ${k===currentSection?'active':''}" data-key="${k}">${escapeHtml(SECTIONS_DATA[k].label)}</button>`
  ).join('');
  document.querySelectorAll('.section-tab').forEach(btn => {
    btn.addEventListener('click', () => loadSection(btn.dataset.key));
  });
}

document.getElementById('product-tabs').addEventListener('click', e => {
  const btn = e.target.closest('.ptab');
  if (!btn) return;
  productFilter = btn.dataset.filter;
  document.querySelectorAll('.ptab').forEach(b => b.classList.toggle('active', b === btn));
  renderTable();
  const visible = filteredProducts();
  if (visible.length && !visible.some(p => p.sku === activeSku)) {
    activeSku = visible[0].sku;
    renderTable();
  }
  renderDetail();
});

document.getElementById('btn-export-pdf').addEventListener('click', () => window.print());

document.getElementById('btn-export-excel').addEventListener('click', () => {
  const list = filteredProducts();
  const rows = [['Producto','SKU','#','Tarea','Estatus','Inicio plan.','Fin plan.',
                  'Inicio real','Fin real','Dur. real (días)','Desvío (días)','Comentarios']];
  list.forEach(p => p.tasks.forEach(t => rows.push([
    p.title, p.sku, t.num, t.name, t.status,
    fmtDisplay(t.plan_start), fmtDisplay(t.plan_end),
    fmtDisplay(t.real_start), fmtDisplay(t.real_end),
    t.real_duration ?? '', t.deviation ?? '', t.comments || ''
  ])));
  const csv = rows.map(r => r.map(v => {
    const s = String(v ?? '');
    return /[",\n]/.test(s) ? '"' + s.replace(/"/g, '""') + '"' : s;
  }).join(',')).join('\r\n');
  const blob = new Blob(['﻿' + csv], {type:'text/csv;charset=utf-8;'});
  const url  = URL.createObjectURL(blob);
  const a    = document.createElement('a');
  a.href = url; a.download = 'cronograma-' + currentSection + '-' + TODAY + '.csv';
  document.body.appendChild(a); a.click(); a.remove();
  URL.revokeObjectURL(url);
});

function filteredProducts() {
  return PRODUCTS.filter(p => productFilter === 'done' ? p.pct_complete === 100 : p.pct_complete < 100);
}

function renderSummary() {
  const total   = PRODUCTS.length;
  const avgPct  = total ? Math.round(PRODUCTS.reduce((a,p) => a + p.pct_complete, 0) / total) : 0;
  const done    = PRODUCTS.filter(p => p.pct_complete === 100).length;
  const atRisk  = PRODUCTS.filter(p => p.max_deviation > 30).length;
  document.getElementById('summary-grid').innerHTML = `
    <div class="kpi"><div class="lbl">Productos en desarrollo</div><div class="val">${total}</div><div class="cnt">Total seguidos en este cronograma</div></div>
    <div class="kpi"><div class="lbl">Progreso promedio</div><div class="val">${avgPct}%</div><div class="cnt">Tareas completadas / total</div></div>
    <div class="kpi"><div class="lbl">Productos completados</div><div class="val">${done}</div><div class="cnt">100% de las tareas finalizadas</div></div>
    <div class="kpi"><div class="lbl">Con desvío &gt; 30 días</div><div class="val" style="color:${atRisk>0?'var(--red)':'var(--accent)'}">${atRisk}</div><div class="cnt">Requieren atención</div></div>`;
}

function renderTable() {
  const tbody = document.getElementById('products-tbody');
  const list  = filteredProducts();
  if (!list.length) {
    tbody.innerHTML = `<tr><td colspan="6" style="text-align:center;color:var(--dim);padding:24px 0;">Sin productos en esta vista.</td></tr>`;
    return;
  }
  tbody.innerHTML = list.map(p => `
    <tr class="product-row ${p.sku===activeSku?'active':''}" data-sku="${escapeHtml(p.sku)}">
      <td><strong>${escapeHtml(p.title)}</strong></td>
      <td><div style="display:flex;align-items:center;gap:8px;">
        <div class="bar-track"><div class="bar-fill" style="width:${p.pct_complete}%"></div></div>
        <span style="font-size:12px;font-weight:700;">${p.pct_complete}%</span>
        ${p.in_progress_tasks > 0 ? `<span class="badge-p status-en-curso">${p.in_progress_tasks} en curso</span>` : ''}
      </div></td>
      <td>${p.completed_tasks}/${p.total_tasks}</td>
      <td>${fmtDisplay(p.plan_start)}</td>
      <td>${fmtDisplay(p.current_end_estimate)}</td>
      <td class="${p.max_deviation > 30 ? 'deviation-warn' : 'deviation-ok'}">${p.max_deviation || 0}</td>
    </tr>`).join('');
  tbody.querySelectorAll('.product-row').forEach(row => {
    row.addEventListener('click', () => {
      activeSku = row.dataset.sku;
      renderTable();
      renderDetail();
    });
  });
}

function renderDetail() {
  const p = PRODUCTS.find(x => x.sku === activeSku);
  if (!p) {
    document.getElementById('detail-title').textContent = '—';
    document.getElementById('detail-sub').textContent   = 'Sin productos en esta vista.';
    document.getElementById('detail-task-table').innerHTML = '';
    document.getElementById('gantt-container').innerHTML   = '';
    return;
  }
  document.getElementById('detail-title').textContent = p.title;
  const shiftNote = p.comite_shift > 0
    ? ` · Desvíos ajustados desde Comité (−${p.comite_shift}d de retraso base)` : '';
  document.getElementById('detail-sub').textContent =
    `${p.completed_tasks}/${p.total_tasks} tareas completadas (${p.pct_complete}%) · Inicio: ${fmtDisplay(p.plan_start)} · Fin estimado: ${fmtDisplay(p.current_end_estimate)}${shiftNote}`;

  // ── Tabla de detalle ───────────────────────────────────────────────────────
  const taskRows = p.tasks.map(t => {
    const dv       = t.adj_deviation !== undefined ? t.adj_deviation : t.deviation;
    const devText  = (dv === null || dv === undefined) ? '—' : dv;
    const durText  = (t.real_duration === null || t.real_duration === undefined) ? '—' : t.real_duration;
    const devClass = (dv || 0) > 14 ? 'deviation-warn' : ((dv || 0) > 0 ? 'deviation-ok' : '');
    const slug     = (t.status || '').toLowerCase().replace(/\s+/g, '-');
    const badge    = `<span class="badge-p status-${slug}">${escapeHtml(t.status || '—')}</span>`;
    return `<tr>
      <td style="text-align:center">${t.num}</td>
      <td>${escapeHtml(t.name)}</td>
      <td>${badge}</td>
      <td>${fmtDisplay(t.plan_start)}</td>
      <td>${fmtDisplay(t.plan_end)}</td>
      <td>${fmtDisplay(t.real_start)}</td>
      <td>${fmtDisplay(t.real_end)}</td>
      <td style="text-align:center">${durText}</td>
      <td style="text-align:center" class="${devClass}">${devText}</td>
    </tr>`;
  }).join('');
  document.getElementById('detail-task-table').innerHTML = `
    <table class="dtable">
      <thead><tr>
        <th>#</th><th>Tarea</th><th>Estatus</th>
        <th>Inicio plan.</th><th>Fin plan.</th>
        <th>Inicio real</th><th>Fin real</th>
        <th>Dur. real (días)</th><th>Desvío (días)</th>
      </tr></thead>
      <tbody>${taskRows}</tbody>
    </table>`;

  // ── Gantt ──────────────────────────────────────────────────────────────────
  const allDates = [];
  p.tasks.forEach(t => {
    [t.plan_start, t.plan_end, t.real_start, t.real_end].forEach(d => {
      if (validDateOrNull(d)) allDates.push(d);
    });
  });
  if (TODAY) allDates.push(TODAY);
  if (!allDates.length) { document.getElementById('gantt-container').innerHTML = ''; return; }

  const minDate   = new Date(Math.min(...allDates.map(d => new Date(d))));
  const maxDate   = new Date(Math.max(...allDates.map(d => new Date(d))));
  const totalSpan = Math.max(1, (maxDate - minDate) / 86400000);

  function pct(dateStr) {
    if (!dateStr) return null;
    return ((new Date(dateStr) - minDate) / 86400000) / totalSpan * 100;
  }

  const todayPct  = pct(TODAY);
  const MONTH_ES  = ['Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic'];
  const monthTicks = [];
  const cursor = new Date(minDate.getFullYear(), minDate.getMonth(), 1);
  while (cursor <= maxDate) {
    const tp = ((cursor - minDate) / 86400000) / totalSpan * 100;
    if (tp >= 0 && tp <= 100)
      monthTicks.push(`<span class="gantt-month-tick" style="left:${tp}%">${MONTH_ES[cursor.getMonth()]} ${cursor.getFullYear()}</span>`);
    cursor.setMonth(cursor.getMonth() + 1);
  }

  const axisRow = `<div class="gantt-group gantt-axis" style="border-bottom:1px solid var(--border);margin-bottom:4px;padding-bottom:4px;">
    <div></div><div></div>
    <div style="padding-left:34px;"><div class="gantt-axis-track">${monthTicks.join('')}</div></div>
    <div></div></div>`;

  document.getElementById('gantt-container').innerHTML = axisRow + p.tasks.map(t => {
    const ps = pct(t.plan_start), pe = pct(t.plan_end);
    const rs = pct(t.real_start), re = pct(t.real_end);
    const todayMarker = todayPct !== null ? `<div class="gantt-today" style="left:${todayPct}%"></div>` : '';
    const dv2      = t.adj_deviation !== undefined ? t.adj_deviation : t.deviation;
    const devClass = (dv2 || 0) > 14 ? 'deviation-warn' : 'deviation-ok';
    const devText  = (dv2 === null || dv2 === undefined) ? '—' : `${dv2}d`;
    const commentEl = t.comments ? `<div class="gantt-comment-inline">💬 ${escapeHtml(t.comments)}</div>` : '';

    const planBar = (ps !== null && pe !== null)
      ? `<div class="gantt-bar" style="left:${ps}%;width:${Math.max(pe-ps,0.6)}%;background:${statusColor(t.status)}"></div>` : '';
    const realBar = (rs !== null && re !== null)
      ? `<div class="gantt-bar-real-solid" style="left:${rs}%;width:${Math.max(re-rs,0.6)}%"></div>` : '';

    return `<div class="gantt-group">
      <div class="gantt-num">${t.num}</div>
      <div class="gantt-name-col">
        <div class="gantt-task" title="${escapeHtml(t.name)}">${escapeHtml(t.name)}</div>
        ${commentEl}
      </div>
      <div class="gantt-bars-col">
        <div class="gantt-bar-row">
          <div class="gantt-bar-label">Plan</div>
          <div class="gantt-track-wrap">${planBar}${todayMarker}</div>
        </div>
        <div class="gantt-bar-row">
          <div class="gantt-bar-label real">Real</div>
          <div class="gantt-track-wrap">${realBar}${todayMarker}</div>
        </div>
      </div>
      <div class="gantt-dev-col">
        <div class="gantt-dev ${devClass}">${devText}</div>
        <div class="gantt-dev-spacer"></div>
      </div>
    </div>`;
  }).join('');
}

function renderAll() {
  const visible = filteredProducts();
  if (!activeSku || !visible.some(p => p.sku === activeSku))
    activeSku = visible.length ? visible[0].sku : null;
  renderSummary();
  renderTable();
  renderDetail();
}

renderSectionTabs();
loadSection(currentSection);
</script>
</body>
</html>
"""


def main():
    try:
        print("Leyendo archivos Excel...")
        sections_data = build_sections_data()

        total = sum(len(s["products"]) for s in sections_data.values())
        if total == 0:
            show_error("No se encontraron productos en los archivos Excel.\nVerifica que los archivos Excel estén en la misma carpeta que este programa.")

        print(f"\nTotal productos: {total}")
        print("Generando dashboard.html...")
        html = build_html(sections_data)
        with open(OUTPUT, "w", encoding="utf-8") as f:
            f.write(html)
        print(f"OK -> {OUTPUT}")

        # Abrir el dashboard automáticamente en el navegador
        import webbrowser
        webbrowser.open(f"file:///{OUTPUT.replace(os.sep, '/')}")
        print("Dashboard abierto en el navegador.")

    except Exception as e:
        show_error(f"Error inesperado:\n{e}")


if __name__ == "__main__":
    main()
