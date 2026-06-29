#!/usr/bin/env python3
"""
gerar_cronograma.py — Cronograma de Nuevos Desarrollos · Cloe
Lee el xlsx de cronogramas (14 productos x 15 tareas), genera cronograma.html
con timeline tipo Gantt, estados, fechas reales vs planeadas y % completado.

Para actualizar: reemplazar el xlsx fuente y volver a correr este script
(o ejecutar ACTUALIZAR CRONOGRAMA CLOE.bat).
"""
import json
import os
import re
from datetime import datetime, date

import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(BASE_DIR))),
    "Copia de Cronogramas_Productos_Nuevos_Cloe_fechas_rev_ok.xlsx",
)
OUTPUT = os.path.join(BASE_DIR, "cronograma.html")

# URL del CSV publicado de Google Sheets ("Archivo > Compartir > Publicar en la web" > CSV).
# Mientras esté vacío, el dashboard usa el snapshot embebido (datos del xlsx en la última generación).
SHEET_CSV_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vSINJkHp80mHYPIF5scyzY3xhmewslj4foW7yShYE_s4GUEtR5jIVm65a3w1Z_0MJR-vcKT6lyVS7C0/pub?gid=528821195&single=true&output=csv"

STATUS_ES = {
    "Completada": "Completada",
    "En curso": "En curso",
    "Pendiente": "Pendiente",
}
STATUS_COLOR = {
    "Completada": "#27AE60",
    "En curso": "#fece08",
    "Pendiente": "#94A3B8",
}


def clean_name(raw):
    name = str(raw or "").strip()
    name = re.sub(r"\s*\(\s*\)\s*", "", name)
    name = name.replace("�", "-").replace("�", "-")
    return name


def fmt_date(v):
    if v is None or not isinstance(v, (datetime, date)):
        return None
    return v.strftime("%Y-%m-%d")


def load_products(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    products = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        title = clean_name(ws["A3"].value) or sheet_name
        start_date = fmt_date(ws["C4"].value)
        today = fmt_date(ws["J4"].value)

        tasks = []
        row = 6
        while True:
            num = ws.cell(row=row, column=1).value
            if num is None:
                break
            task_name = clean_name(ws.cell(row=row, column=2).value)
            duration = ws.cell(row=row, column=3).value
            depends_on = ws.cell(row=row, column=4).value
            dep_type = ws.cell(row=row, column=5).value
            plan_start = fmt_date(ws.cell(row=row, column=6).value)
            plan_end = fmt_date(ws.cell(row=row, column=7).value)
            status = str(ws.cell(row=row, column=8).value or "Pendiente").strip()
            real_start = fmt_date(ws.cell(row=row, column=9).value)
            real_end = fmt_date(ws.cell(row=row, column=10).value)
            real_dur = ws.cell(row=row, column=11).value
            deviation = ws.cell(row=row, column=12).value
            comments = str(ws.cell(row=row, column=13).value or "").strip()

            tasks.append({
                "num": int(num),
                "name": task_name,
                "duration": duration,
                "depends_on": depends_on,
                "dep_type": dep_type,
                "plan_start": plan_start,
                "plan_end": plan_end,
                "status": status if status in STATUS_COLOR else "Pendiente",
                "real_start": real_start,
                "real_end": real_end,
                "real_duration": real_dur,
                "deviation": deviation,
                "comments": comments,
            })
            row += 1

        total = len(tasks)
        completed = sum(1 for t in tasks if t["status"] == "Completada")
        in_progress = sum(1 for t in tasks if t["status"] == "En curso")
        pct = round(100 * completed / total) if total else 0

        all_plan_starts = [t["plan_start"] for t in tasks if t["plan_start"]]
        all_plan_ends = [t["plan_end"] for t in tasks if t["plan_end"]]
        all_real_ends = [t["real_end"] for t in tasks if t["real_end"]]

        max_deviation = max(
            [t["deviation"] for t in tasks if isinstance(t["deviation"], (int, float))],
            default=0,
        )

        # "Inicio" se referencia a la tarea "Reunión con Comité" (#4): fecha real si existe,
        # si no la planeada; si esa tarea no existe, se usa la fecha planeada más temprana.
        comite_task = next((t for t in tasks if t["num"] == 4), None)
        comite_start = (comite_task and (comite_task["real_end"] or comite_task["real_start"] or comite_task["plan_end"] or comite_task["plan_start"])) or None
        display_start = comite_start or (min(all_plan_starts) if all_plan_starts else None)

        products.append({
            "sku": sheet_name,
            "title": title,
            "start_date": start_date,
            "today": today,
            "tasks": tasks,
            "total_tasks": total,
            "completed_tasks": completed,
            "in_progress_tasks": in_progress,
            "pct_complete": pct,
            "plan_start": display_start,
            "plan_end": max(all_plan_ends) if all_plan_ends else None,
            "current_end_estimate": max(all_real_ends) if all_real_ends else (max(all_plan_ends) if all_plan_ends else None),
            "max_deviation": max_deviation,
        })

    return products


def build_html(products):
    today = products[0]["today"] if products else date.today().strftime("%Y-%m-%d")
    data_json = json.dumps(products, ensure_ascii=False)
    generated_at = datetime.now().strftime("%d/%m/%Y %H:%M")

    html = HTML_TEMPLATE.replace("__DATA_JSON__", data_json)
    html = html.replace("__GENERATED_AT__", generated_at)
    html = html.replace("__TODAY__", today or "")
    html = html.replace("__N_PRODUCTS__", str(len(products)))
    html = html.replace("__SHEET_CSV_URL__", SHEET_CSV_URL)
    return html


HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<meta name="robots" content="noindex, nofollow">
<title>Cronograma Nuevos Desarrollos · Cloe</title>
<style>
:root{--bg:#0a0e1a;--card:#111827;--border:#1e293b;--text:#e2e8f0;--dim:#64748b;
  --accent:#f59e0b;--green:#10b981;--red:#ef4444;--blue:#3b82f6;--purple:#8b5cf6;
  --c0:#ef4444;--c1:#f97316;--c2:#eab308;--c3:#a855f7;--c4:#64748b;}
*{margin:0;padding:0;box-sizing:border-box}
body{background:var(--bg);color:var(--text);font-family:'Segoe UI',system-ui,sans-serif;font-size:13px;line-height:1.5}
.header{background:linear-gradient(135deg,#111827 0%,#1e293b 100%);padding:14px 24px;border-bottom:2px solid var(--accent);display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:8px;}
.logo{display:flex;align-items:center;gap:10px;}
.logo-text .brand{font-size:17px;font-weight:700;color:#fff;display:block;}
.logo-text .sub{font-size:11px;color:var(--dim);}
.header-right{text-align:right;font-size:11px;color:var(--dim);}
.kpi-row{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;padding:16px 24px;}
.kpi{background:var(--card);border:1px solid var(--border);border-radius:10px;padding:14px 18px;text-align:center;}
.kpi .lbl{font-size:10px;color:var(--dim);text-transform:uppercase;letter-spacing:.5px;}
.kpi .val{font-size:20px;font-weight:800;color:var(--accent);margin:4px 0;}
.kpi .cnt{font-size:11px;color:var(--dim);}
.content{padding:16px 24px;max-width:1600px;margin:0 auto;}
.card{background:var(--card);border:1px solid var(--border);border-radius:10px;padding:18px;margin-bottom:16px;}
.section-hdr{display:flex;align-items:center;justify-content:space-between;gap:8px;margin-bottom:10px;flex-wrap:wrap;}
.section-hdr h2{font-size:13px;font-weight:700;color:var(--accent);text-transform:uppercase;letter-spacing:.4px;}
.ptabs{display:flex;gap:6px;}
.ptab{padding:5px 14px;border-radius:6px;font-size:11px;font-weight:600;cursor:pointer;background:#1e293b;border:1px solid var(--border);color:var(--dim);transition:all .15s;}
.ptab:hover{color:var(--text);}
.ptab.active{background:#f59e0b22;color:var(--accent);border-color:var(--accent);}
#sync-banner{background:#f59e0b22;color:var(--accent);border:1px solid #f59e0b55;border-radius:8px;padding:10px 14px;font-size:12px;margin-bottom:14px;}
.tbl-wrap{overflow-x:auto;}
table.dt{width:100%;border-collapse:collapse;font-size:11px;}
table.dt th{text-align:left;padding:5px 8px;border-bottom:2px solid var(--accent);color:var(--accent);font-weight:700;font-size:10px;text-transform:uppercase;white-space:nowrap;}
table.dt td{padding:6px 8px;border-bottom:1px solid var(--border);white-space:nowrap;}
table.dt tr.product-row{cursor:pointer;}
table.dt tr.product-row:hover td{background:#1e293b55;}
table.dt tr.product-row.active td{background:#f59e0b18;}
.bar-track{background:#1e293b;border-radius:4px;height:8px;width:140px;overflow:hidden;display:inline-block;vertical-align:middle;}
.bar-fill{height:100%;border-radius:4px;background:var(--green);}
.badge-p{display:inline-block;padding:2px 8px;border-radius:20px;font-size:11px;font-weight:700;}
.status-pendiente{background:#64748b22;color:var(--dim);}
.status-curso{background:#f59e0b22;color:var(--accent);}
.status-completada{background:#10b98122;color:var(--green);}
.deviation-ok{color:var(--green);font-weight:700;}
.deviation-warn{color:var(--red);font-weight:700;}

.detail-title{font-size:15px;font-weight:800;color:#fff;}
.detail-sub{font-size:11px;color:var(--dim);margin-top:2px;}
.legend{display:flex;gap:16px;flex-wrap:wrap;margin-bottom:14px;font-size:11px;color:var(--dim);}
.legend-item{display:flex;align-items:center;gap:6px;}
.legend-dot{width:10px;height:10px;border-radius:3px;display:inline-block;}

.gantt{display:flex;flex-direction:column;gap:6px;}
.gantt-row{display:grid;grid-template-columns:26px 250px 1fr 70px;align-items:center;gap:10px;}
.gantt-num{font-size:11px;color:var(--dim);text-align:right;}
.gantt-task{font-size:12px;font-weight:500;color:var(--text);}
.gantt-track-wrap{position:relative;height:20px;background:#1e293b;border-radius:5px;overflow:hidden;}
.gantt-bar{position:absolute;top:2px;bottom:2px;border-radius:4px;opacity:.92;}
.gantt-bar-real{position:absolute;top:2px;bottom:2px;border-radius:4px;border:2px solid rgba(255,255,255,.35);background:transparent;}
.gantt-today{position:absolute;top:-2px;bottom:-2px;width:2px;background:var(--accent);}
.gantt-dev{font-size:11px;text-align:right;}
.gantt-comment{grid-column:2/4;font-size:11px;color:var(--dim);font-style:italic;padding:0 0 4px 0;}

footer{text-align:center;color:var(--dim);font-size:11px;padding:16px;border-top:1px solid var(--border);margin-top:8px;}

@media(max-width:768px){
  .kpi-row{grid-template-columns:repeat(2,1fr);}
  .gantt-row{grid-template-columns:22px 1fr;grid-template-areas:"num task" "bar bar" "dev dev";}
  .gantt-task{grid-area:task;}
  .gantt-num{grid-area:num;}
  .gantt-track-wrap{grid-area:bar;}
  .gantt-dev{grid-area:dev;text-align:left;}
}
</style>
</head>
<body>

<div class="header">
  <div class="logo">
    <svg viewBox="0 0 48 48" width="40" height="40" xmlns="http://www.w3.org/2000/svg">
      <rect width="48" height="48" rx="8" fill="#1B2A6B"/>
      <path d="M10 34 Q16 10 24 24 Q32 38 38 14" stroke="#FFC72C" stroke-width="3.5" fill="none" stroke-linecap="round"/>
      <circle cx="38" cy="14" r="2.5" fill="#FFC72C"/>
    </svg>
    <div class="logo-text">
      <span class="brand">Travel Blue</span>
      <span class="sub">Cronograma · Nuevos Desarrollos Cloe</span>
    </div>
  </div>
  <div class="header-right">__N_PRODUCTS__ productos · Generado el __GENERATED_AT__</div>
</div>

<div class="kpi-row" id="summary-grid"></div>

<div class="content">
  <div id="sync-banner" style="display:none;"></div>

  <div class="card">
    <div class="section-hdr">
      <h2>📋 Resumen por producto</h2>
      <div class="ptabs" id="product-tabs">
        <button class="ptab active" data-filter="active">En desarrollo</button>
        <button class="ptab" data-filter="done">Completados</button>
      </div>
    </div>
    <div class="tbl-wrap">
      <table class="dt">
        <thead>
          <tr>
            <th>Producto (SKU)</th>
            <th>Progreso</th>
            <th>Tareas</th>
            <th>Inicio plan.</th>
            <th>Fin estimado</th>
            <th>Desvío máx.</th>
          </tr>
        </thead>
        <tbody id="products-tbody"></tbody>
      </table>
    </div>
  </div>

  <div class="card">
    <div class="section-hdr"><h2>📅 Cronograma detallado</h2></div>
    <div class="legend">
      <span class="legend-item"><span class="legend-dot" style="background:#10b981"></span> Completada</span>
      <span class="legend-item"><span class="legend-dot" style="background:#f59e0b"></span> En curso</span>
      <span class="legend-item"><span class="legend-dot" style="background:#64748b"></span> Pendiente</span>
      <span class="legend-item"><span class="legend-dot" style="border:2px solid rgba(255,255,255,.5); background:transparent"></span> Fechas reales</span>
      <span class="legend-item"><span class="legend-dot" style="background:#f59e0b;width:2px;border-radius:0"></span> Hoy (__TODAY__)</span>
    </div>
    <div class="detail-title" id="detail-title">—</div>
    <div class="detail-sub" id="detail-sub" style="margin-bottom:14px;">—</div>
    <div class="gantt" id="gantt-container"></div>
  </div>
</div>

<footer>Travel Blue · Cronograma compartido con Cloe · Última actualización: __GENERATED_AT__</footer>

<script>
const FALLBACK_PRODUCTS = __DATA_JSON__;
const FALLBACK_TODAY = "__TODAY__";
const SHEET_CSV_URL = "__SHEET_CSV_URL__";

let PRODUCTS = FALLBACK_PRODUCTS;
let TODAY = FALLBACK_TODAY;

function showBanner(msg) {
  const b = document.getElementById('sync-banner');
  if (!msg) { b.style.display = 'none'; return; }
  b.style.display = 'block';
  b.textContent = msg;
}

function parseCsv(text) {
  const lines = text.trim().split(/\r?\n/);
  const headers = splitCsvLine(lines[0]);
  return lines.slice(1).filter(l => l.trim()).map(line => {
    const cells = splitCsvLine(line);
    const obj = {};
    headers.forEach((h, i) => { obj[h.trim()] = (cells[i] || '').trim(); });
    return obj;
  });
}
function splitCsvLine(line) {
  const out = []; let cur = ''; let inQ = false;
  for (let i = 0; i < line.length; i++) {
    const c = line[i];
    if (inQ) {
      if (c === '"' && line[i+1] === '"') { cur += '"'; i++; }
      else if (c === '"') { inQ = false; }
      else { cur += c; }
    } else {
      if (c === '"') inQ = true;
      else if (c === ',') { out.push(cur); cur = ''; }
      else cur += c;
    }
  }
  out.push(cur);
  return out;
}

function escapeHtml(s) {
  return String(s == null ? '' : s).replace(/[&<>"']/g, c => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c]));
}

function numOrNull(v) {
  if (v === undefined || v === null || v === '') return null;
  const n = Number(v);
  return isNaN(n) ? null : n;
}

function validDateOrNull(v) {
  if (!v || !/^\d{4}-\d{2}-\d{2}$/.test(v)) return null;
  return v;
}

function buildProductsFromRows(rows) {
  const bySku = new Map();
  rows.forEach(r => {
    const sku = r.SKU;
    if (!sku) return;
    if (!bySku.has(sku)) {
      bySku.set(sku, { sku, title: r.Producto || sku, tasks: [] });
    }
    bySku.get(sku).tasks.push({
      num: numOrNull(r['#']),
      name: r.Tarea || '',
      duration: numOrNull(r.Duracion_dias),
      depends_on: r.Depende_de || null,
      dep_type: r.Tipo_dep || null,
      plan_start: validDateOrNull(r.Inicio_plan),
      plan_end: validDateOrNull(r.Fin_plan),
      status: ['Completada', 'En curso', 'Pendiente'].includes(r.Estatus) ? r.Estatus : 'Pendiente',
      real_start: validDateOrNull(r.Inicio_real),
      real_end: validDateOrNull(r.Fin_real),
      real_duration: numOrNull(r.Dur_real_dias),
      deviation: numOrNull(r.Desvio_dias),
      comments: r.Comentarios || '',
    });
  });
  const products = [];
  bySku.forEach(p => {
    const total = p.tasks.length;
    const completed = p.tasks.filter(t => t.status === 'Completada').length;
    const inProgress = p.tasks.filter(t => t.status === 'En curso').length;
    const planStarts = p.tasks.map(t => t.plan_start).filter(Boolean);
    const planEnds = p.tasks.map(t => t.plan_end).filter(Boolean);
    const realEnds = p.tasks.map(t => t.real_end).filter(Boolean);
    const deviations = p.tasks.map(t => t.deviation).filter(d => typeof d === 'number');
    products.push({
      sku: p.sku,
      title: p.title,
      tasks: p.tasks,
      total_tasks: total,
      completed_tasks: completed,
      in_progress_tasks: inProgress,
      pct_complete: total ? Math.round(100 * completed / total) : 0,
      plan_start: planStarts.length ? planStarts.sort()[0] : null,
      plan_end: planEnds.length ? planEnds.sort().slice(-1)[0] : null,
      current_end_estimate: realEnds.length ? realEnds.sort().slice(-1)[0] : (planEnds.length ? planEnds.sort().slice(-1)[0] : null),
      max_deviation: deviations.length ? Math.max(...deviations) : 0,
    });
  });
  return products;
}

async function loadData() {
  if (!SHEET_CSV_URL) {
    showBanner('Mostrando datos de la última actualización del archivo (__GENERATED_AT__). Conecta la planilla de Google Sheets para que esto se actualice solo.');
    renderAll();
    return;
  }
  try {
    const res = await fetch(SHEET_CSV_URL + (SHEET_CSV_URL.includes('?') ? '&' : '?') + 'cb=' + Date.now());
    if (!res.ok) throw new Error('HTTP ' + res.status);
    const text = await res.text();
    const rows = parseCsv(text);
    const live = buildProductsFromRows(rows);
    if (!live.length) throw new Error('CSV vacío');
    PRODUCTS = live;
    TODAY = new Date().toISOString().slice(0, 10);
    showBanner('');
  } catch (e) {
    showBanner('No se pudo conectar a la planilla en vivo — mostrando la última actualización conocida.');
  }
  renderAll();
}

function toggleTheme() {
  const isDark = document.body.classList.contains('dark');
  if (isDark) { document.body.classList.remove('dark'); localStorage.setItem('tb_theme','light'); }
  else { document.body.classList.add('dark'); localStorage.setItem('tb_theme','dark'); }
}
(function(){
  var theme = localStorage.getItem('tb_theme');
  if (theme === 'dark') document.body.classList.add('dark');
})();

function statusClass(s) {
  if (s === 'Completada') return 'status-completada';
  if (s === 'En curso') return 'status-curso';
  return 'status-pendiente';
}
function statusColor(s) {
  if (s === 'Completada') return '#27AE60';
  if (s === 'En curso') return '#fece08';
  return '#94A3B8';
}

let activeSku = PRODUCTS.length ? PRODUCTS[0].sku : null;
let productFilter = 'active';

document.getElementById('product-tabs').addEventListener('click', (e) => {
  const btn = e.target.closest('.ptab');
  if (!btn) return;
  productFilter = btn.dataset.filter;
  document.querySelectorAll('.ptab').forEach(b => b.classList.toggle('active', b === btn));
  renderTable();
  const visible = filteredProducts();
  if (visible.length && !visible.some(p => p.sku === activeSku)) {
    activeSku = visible[0].sku;
    renderTable();
  }
  renderDetail();
});

function filteredProducts() {
  return PRODUCTS.filter(p => productFilter === 'done' ? p.pct_complete === 100 : p.pct_complete < 100);
}

function renderSummary() {
  const totalProducts = PRODUCTS.length;
  const avgPct = totalProducts ? Math.round(PRODUCTS.reduce((a,p)=>a+p.pct_complete,0)/totalProducts) : 0;
  const completedProducts = PRODUCTS.filter(p => p.pct_complete === 100).length;
  const atRisk = PRODUCTS.filter(p => p.max_deviation > 30).length;
  document.getElementById('summary-grid').innerHTML = `
    <div class="kpi"><div class="lbl">Productos en desarrollo</div><div class="val">${totalProducts}</div><div class="cnt">Total seguidos en este cronograma</div></div>
    <div class="kpi"><div class="lbl">Progreso promedio</div><div class="val">${avgPct}%</div><div class="cnt">Tareas completadas / total</div></div>
    <div class="kpi"><div class="lbl">Productos completados</div><div class="val">${completedProducts}</div><div class="cnt">100% de las tareas finalizadas</div></div>
    <div class="kpi"><div class="lbl" >Con desvío &gt; 30 días</div><div class="val" style="color:${atRisk>0?'var(--red)':'var(--accent)'}">${atRisk}</div><div class="cnt">Requieren atención</div></div>
  `;
}

function renderTable() {
  const tbody = document.getElementById('products-tbody');
  const list = filteredProducts();
  if (!list.length) {
    tbody.innerHTML = `<tr><td colspan="6" style="text-align:center;color:var(--dim);padding:24px 0;">Sin productos en esta vista.</td></tr>`;
    return;
  }
  tbody.innerHTML = list.map(p => `
    <tr class="product-row ${p.sku===activeSku?'active':''}" data-sku="${p.sku}">
      <td><strong>${p.title}</strong></td>
      <td>
        <div style="display:flex;align-items:center;gap:8px;">
          <div class="bar-track"><div class="bar-fill" style="width:${p.pct_complete}%"></div></div>
          <span style="font-size:12px;font-weight:700;">${p.pct_complete}%</span>
          ${p.in_progress_tasks > 0 ? `<span class="badge-p status-curso">${p.in_progress_tasks} en curso</span>` : ''}
        </div>
      </td>
      <td>${p.completed_tasks}/${p.total_tasks}</td>
      <td>${p.plan_start || '—'}</td>
      <td>${p.current_end_estimate || '—'}</td>
      <td class="${p.max_deviation > 30 ? 'deviation-warn' : 'deviation-ok'}">${p.max_deviation || 0}</td>
    </tr>
  `).join('');
  tbody.querySelectorAll('.product-row').forEach(row => {
    row.addEventListener('click', () => { activeSku = row.dataset.sku; renderTable(); renderDetail(); });
  });
}

function daysBetween(a, b) {
  return (new Date(b) - new Date(a)) / 86400000;
}

function renderDetail() {
  const p = PRODUCTS.find(x => x.sku === activeSku);
  if (!p) {
    document.getElementById('detail-title').textContent = '—';
    document.getElementById('detail-sub').textContent = 'Sin productos en esta vista.';
    document.getElementById('gantt-container').innerHTML = '';
    return;
  }
  document.getElementById('detail-title').textContent = p.title;
  document.getElementById('detail-sub').textContent =
    `${p.completed_tasks}/${p.total_tasks} tareas completadas (${p.pct_complete}%) · Inicio: ${p.plan_start || '—'} · Fin estimado: ${p.current_end_estimate || '—'}`;

  const allDates = [];
  p.tasks.forEach(t => {
    [t.plan_start, t.plan_end, t.real_start, t.real_end].forEach(d => { if (validDateOrNull(d)) allDates.push(d); });
  });
  if (TODAY) allDates.push(TODAY);
  const minDate = new Date(Math.min(...allDates.map(d => new Date(d))));
  const maxDate = new Date(Math.max(...allDates.map(d => new Date(d))));
  const totalSpan = Math.max(1, (maxDate - minDate) / 86400000);

  function pct(dateStr) {
    if (!dateStr) return null;
    return ((new Date(dateStr) - minDate) / 86400000) / totalSpan * 100;
  }

  const todayPct = pct(TODAY);

  const container = document.getElementById('gantt-container');
  container.innerHTML = p.tasks.map(t => {
    const ps = pct(t.plan_start), pe = pct(t.plan_end);
    const rs = pct(t.real_start), re = pct(t.real_end);
    let planBar = '';
    if (ps !== null && pe !== null) {
      planBar = `<div class="gantt-bar" style="left:${ps}%;width:${Math.max(pe-ps,0.6)}%;background:${statusColor(t.status)}"></div>`;
    }
    let realBar = '';
    if (rs !== null && re !== null) {
      realBar = `<div class="gantt-bar-real" style="left:${rs}%;width:${Math.max(re-rs,0.6)}%;"></div>`;
    }
    const todayMarker = todayPct !== null ? `<div class="gantt-today" style="left:${todayPct}%"></div>` : '';
    const devClass = (t.deviation || 0) > 14 ? 'deviation-warn' : 'deviation-ok';
    const devText = (t.deviation === null || t.deviation === undefined) ? '—' : `${t.deviation}d`;
    const commentRow = t.comments ? `<div class="gantt-comment">💬 ${escapeHtml(t.comments)}</div>` : '';
    return `
      <div class="gantt-row">
        <div class="gantt-num">${t.num}</div>
        <div class="gantt-task" title="${escapeHtml(t.name)}">${escapeHtml(t.name)}</div>
        <div class="gantt-track-wrap">${planBar}${realBar}${todayMarker}</div>
        <div class="gantt-dev ${devClass}">${devText}</div>
      </div>
      ${commentRow}
    `;
  }).join('');
}

function renderAll() {
  const visible = filteredProducts();
  if (activeSku === null && visible.length) activeSku = visible[0].sku;
  if (activeSku && !visible.some(p => p.sku === activeSku)) activeSku = visible.length ? visible[0].sku : null;
  renderSummary();
  renderTable();
  renderDetail();
}

loadData();
</script>
</body>
</html>
"""


def main():
    products = load_products(SRC)
    html = build_html(products)
    with open(OUTPUT, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"OK: {len(products)} productos -> {OUTPUT}")


if __name__ == "__main__":
    main()
